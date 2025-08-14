from fastapi import FastAPI, Request, Form, File, UploadFile, Response
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from contextlib import asynccontextmanager
import io
import json
import os
from datetime import datetime
from openpyxl import load_workbook
from pydantic import BaseModel
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()
from models import CheckinResponse, ImportResponse, User, DeleteResponse, CreateUserResponse, Settings, SettingsUpdate, SettingsResponse
from database import init_db, get_user_by_employee_id, create_checkin, get_checkin_history, create_users_batch, get_all_users, delete_all_users, create_single_user, search_users, get_tables_with_users, get_export_data, clear_checkin_history, checkout_user, get_settings, update_settings, has_admin_user, create_initial_admin_if_needed, create_auth_user, authenticate_user, get_auth_user, get_all_auth_users, delete_auth_user, create_session, delete_session, cleanup_expired_sessions
from auth import AuthMiddleware

# Auth models
class LoginRequest(BaseModel):
    username: str
    password: str

class AuthUserRequest(BaseModel):
    username: str
    password: str
    is_admin: bool = False

@asynccontextmanager
async def lifespan(_: FastAPI):
    init_db()
    # Create initial admin from environment variables if needed
    create_initial_admin_if_needed()
    # Clean up expired sessions on startup
    cleanup_expired_sessions()
    yield

app = FastAPI(title="RFID Checkin Station", lifespan=lifespan)

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

templates = Jinja2Templates(directory="templates")

# Authentication routes
@app.get("/auth/login", response_class=HTMLResponse)
async def login_page(request: Request):
    # If already logged in, redirect to home
    if AuthMiddleware.is_authenticated(request):
        return RedirectResponse(url="/", status_code=302)
    
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/auth/login")
async def login(http_request: Request, request: LoginRequest):
    if not request.username or not request.password:
        return {"success": False, "message": "Username and password required"}
    
    # Authenticate user
    user = authenticate_user(request.username, request.password)
    if not user:
        return {"success": False, "message": "Invalid username or password"}
    
    # Create session
    session_id = create_session(user["username"])
    
    # Create response with session cookie

    response_data = {"success": True, "message": "Login successful"}
    resp = Response(content=json.dumps(response_data), media_type="application/json")
    
    # Use secure cookies only in production (when HTTPS is available)
    # For local development, allow HTTP cookies
    is_production = os.getenv("ENVIRONMENT", "development").lower() == "production"
    
    # Check for common misconfiguration: production mode without HTTPS
    if is_production and not http_request.url.scheme == "https":
        return {
            "success": False, 
            "message": "Configuration Error: ENVIRONMENT is set to 'production' but you're not using HTTPS. Either use HTTPS or set ENVIRONMENT=development for local testing."
        }
    
    resp.set_cookie(
        key="session_id",
        value=session_id,
        max_age=30 * 24 * 60 * 60,  # 30 days
        httponly=True,
        secure=is_production,  # True in production, False in development
        samesite="lax"
    )
    return resp

@app.post("/auth/logout")
async def logout(request: Request):
    session_id = request.cookies.get("session_id")
    if session_id:
        delete_session(session_id)
    
    response = RedirectResponse(url="/auth/login", status_code=302)
    response.delete_cookie("session_id")
    return response

# Main routes (now protected)
@app.get("/", response_class=HTMLResponse)
async def checkin_page(request: Request):
    # Check authentication
    if not AuthMiddleware.is_authenticated(request):
        return RedirectResponse(url="/auth/login", status_code=302)
    
    # Check if user is admin to show/hide admin link
    show_admin_link = AuthMiddleware.is_admin(request)
    
    settings = get_settings()
    return templates.TemplateResponse("checkin.html", {
        "request": request, 
        "settings": settings,
        "show_admin_link": show_admin_link
    })

@app.get("/preview", response_class=HTMLResponse)
async def checkin_preview(request: Request, demo_result: bool = False):
    settings = get_settings()
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return templates.TemplateResponse("checkin.html", {"request": request, "settings": settings, "current_time": current_time, "preview_mode": True, "demo_result": demo_result, "show_admin_link": False, "show_footer": False})

@app.post("/checkin", response_model=CheckinResponse)
async def checkin(badge_id: str = Form(...)):
    user = get_user_by_employee_id(badge_id)
    
    if user:
        success = create_checkin(badge_id)
        if success:
            return CheckinResponse(
                success=True,
                name=f"{user.first_name} {user.last_name}",
                table_number=user.table_number,
                time=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
        else:
            return CheckinResponse(success=False, message="Checkin failed")
    else:
        return CheckinResponse(success=False, message="Badge not found. Please see check-in attendant.")

@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    if not AuthMiddleware.is_authenticated(request):
        return RedirectResponse(url="/auth/login", status_code=302)
    
    # Require admin privileges
    if not AuthMiddleware.is_admin(request):
        # Redirect to main page with error or show access denied
        return RedirectResponse(url="/?error=admin_required", status_code=302)
    
    return templates.TemplateResponse("admin.html", {"request": request, "show_admin_link": False})

# Login Users Management (Admin Only)
@app.get("/admin/auth-users")
async def get_auth_users(request: Request):
    AuthMiddleware.require_admin(request)
    return get_all_auth_users()

@app.post("/admin/auth-users")
async def create_auth_user_endpoint(request: Request, user_request: AuthUserRequest):
    AuthMiddleware.require_admin(request)
    
    if not user_request.username or not user_request.password:
        return {"success": False, "message": "Username and password are required"}
    
    if len(user_request.password) < 6:
        return {"success": False, "message": "Password must be at least 6 characters"}
    
    if create_auth_user(user_request.username, user_request.password, user_request.is_admin):
        return {"success": True, "message": "Login user created successfully"}
    else:
        return {"success": False, "message": "Username already exists"}

@app.delete("/admin/auth-users/{username}")
async def delete_auth_user_endpoint(request: Request, username: str):
    AuthMiddleware.require_admin(request)
    
    if delete_auth_user(username):
        return {"success": True, "message": "Login user deleted successfully"}
    else:
        return {"success": False, "message": "Cannot delete user (user not found or last admin)"}

@app.get("/admin/history")
async def get_history(request: Request, search: str = ""):
    AuthMiddleware.require_admin(request)
    return get_checkin_history(search)

@app.get("/admin/users")
async def get_users(request: Request, search: str = ""):
    AuthMiddleware.require_admin(request)
    if search:
        return search_users(search)
    return get_all_users()

@app.get("/admin/tables")
async def get_tables(request: Request, search: str = ""):
    AuthMiddleware.require_admin(request)
    return get_tables_with_users(search)

@app.get("/admin/export")
async def export_xlsx(request: Request):
    AuthMiddleware.require_admin(request)
    from openpyxl import Workbook
    
    export_data = get_export_data()
    
    # Create a new workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Checkin Data"
    
    current_row = 1
    
    # Section 1: Users with checkins
    if export_data['with_checkins']:
        # Add section header
        worksheet.cell(row=current_row, column=1, value="USERS WITH CHECKINS")
        worksheet.cell(row=current_row, column=1).font = worksheet.cell(row=current_row, column=1).font.copy(bold=True)
        current_row += 1
        
        # Add headers for checkins section
        headers = ["First Name", "Last Name", "Employee ID", "Table Number", "Checkin Time"]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=current_row, column=col, value=header)
            worksheet.cell(row=current_row, column=col).font = worksheet.cell(row=current_row, column=col).font.copy(bold=True)
        current_row += 1
        
        # Add checkin data
        for record in export_data['with_checkins']:
            worksheet.cell(row=current_row, column=1, value=record.first_name)
            worksheet.cell(row=current_row, column=2, value=record.last_name)
            worksheet.cell(row=current_row, column=3, value=record.employee_id)
            worksheet.cell(row=current_row, column=4, value=record.table_number)
            worksheet.cell(row=current_row, column=5, value=record.checkin_time)
            current_row += 1
    
    # Section 2: Users without checkins
    if export_data['without_checkins']:
        # Add spacing between sections
        current_row += 1
        
        # Add section header
        worksheet.cell(row=current_row, column=1, value="USERS WITHOUT CHECKINS")
        worksheet.cell(row=current_row, column=1).font = worksheet.cell(row=current_row, column=1).font.copy(bold=True)
        current_row += 1
        
        # Add headers for no-checkins section
        headers = ["First Name", "Last Name", "Employee ID", "Table Number", "Status"]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=current_row, column=col, value=header)
            worksheet.cell(row=current_row, column=col).font = worksheet.cell(row=current_row, column=col).font.copy(bold=True)
        current_row += 1
        
        # Add users without checkins
        for user in export_data['without_checkins']:
            worksheet.cell(row=current_row, column=1, value=user.first_name)
            worksheet.cell(row=current_row, column=2, value=user.last_name)
            worksheet.cell(row=current_row, column=3, value=user.employee_id)
            worksheet.cell(row=current_row, column=4, value=user.table_number)
            worksheet.cell(row=current_row, column=5, value="No Checkin")
            current_row += 1
    
    # Save to BytesIO
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    def iter_xlsx():
        yield output.read()
    
    return StreamingResponse(
        iter_xlsx(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=checkin_data.xlsx"}
    )

@app.post("/admin/import", response_model=ImportResponse)
async def import_users(request: Request, file: UploadFile = File(...)):
    AuthMiddleware.require_admin(request)
    if not file.filename:
        return ImportResponse(success=False, message="Please upload a file")
    
    # Check file extension
    filename = file.filename.lower()
    if not filename.endswith('.xlsx'):
        return ImportResponse(success=False, message="Please upload an Excel (.xlsx) file")
    
    content = await file.read()
    users = []
    errors = []
    
    try:
        # Load Excel file
        workbook = load_workbook(io.BytesIO(content))
        worksheet = workbook.active
        
        # Read header row to map column positions
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        if not header_row:
            return ImportResponse(success=False, message="Excel file appears to be empty")
        
        # Create header mapping (case-insensitive)
        header_map = {}
        for col_idx, header in enumerate(header_row):
            if header:
                header_lower = str(header).lower().strip()
                header_map[header_lower] = col_idx
        
        # Required headers (flexible matching)
        required_fields = {
            'first_name': ['first name', 'firstname', 'first', 'fname'],
            'last_name': ['last name', 'lastname', 'last', 'lname', 'surname'],
            'employee_id': ['employee id', 'employeeid', 'employee', 'id', 'badge', 'badge id', 'employe id', 'employeid', 'emp id', 'emp_id'],
            'table_number': ['table number', 'tablenumber', 'table', 'table_number', 'table num']
        }
        
        # Find column indices for each required field
        column_indices = {}
        for field, possible_headers in required_fields.items():
            found = False
            for possible_header in possible_headers:
                if possible_header in header_map:
                    column_indices[field] = header_map[possible_header]
                    found = True
                    break
            if not found:
                missing_headers = ', '.join(possible_headers[:3])  # Show first 3 options
                return ImportResponse(
                    success=False, 
                    message=f"Missing required column for {field}. Expected one of: {missing_headers}"
                )
        
        # Process data rows
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            try:
                # Extract values using column mapping
                first_name_val = row[column_indices['first_name']] if column_indices['first_name'] < len(row) else None
                last_name_val = row[column_indices['last_name']] if column_indices['last_name'] < len(row) else None
                employee_id_val = row[column_indices['employee_id']] if column_indices['employee_id'] < len(row) else None
                table_number_val = row[column_indices['table_number']] if column_indices['table_number'] < len(row) else None
                
                # Validate required fields are not empty
                if not first_name_val or str(first_name_val).strip() == '':
                    errors.append(f"Row {row_num}: Missing first name")
                    continue
                if not last_name_val or str(last_name_val).strip() == '':
                    errors.append(f"Row {row_num}: Missing last name")
                    continue
                if not employee_id_val or str(employee_id_val).strip() == '':
                    errors.append(f"Row {row_num}: Missing employee ID")
                    continue
                if not table_number_val:
                    errors.append(f"Row {row_num}: Missing table number")
                    continue
                
                # Create user object
                user = User(
                    first_name=str(first_name_val).strip(),
                    last_name=str(last_name_val).strip(),
                    employee_id=str(employee_id_val).strip(),
                    table_number=int(table_number_val)
                )
                users.append(user)
                
            except (ValueError, TypeError) as e:
                errors.append(f"Row {row_num}: {str(e)}")
            except Exception as e:
                errors.append(f"Row {row_num}: Unexpected error - {str(e)}")
    
    except Exception as e:
        return ImportResponse(success=False, message=f"Error reading Excel file: {str(e)}")
    
    if users:
        imported, db_errors = create_users_batch(users)
        errors.extend(db_errors)
        return ImportResponse(success=True, imported=imported, errors=errors)
    else:
        return ImportResponse(success=False, message="No valid users found", errors=errors)

@app.delete("/admin/users", response_model=DeleteResponse)
async def delete_all_users_endpoint(request: Request):
    AuthMiddleware.require_admin(request)
    try:
        deleted_count = delete_all_users()
        return DeleteResponse(
            success=True,
            deleted=deleted_count,
            message=f"Successfully deleted {deleted_count} users"
        )
    except Exception as e:
        return DeleteResponse(
            success=False,
            message=f"Error deleting users: {str(e)}"
        )

@app.post("/admin/users", response_model=CreateUserResponse)
async def create_user_endpoint(request: Request, user: User):
    AuthMiddleware.require_admin(request)
    try:
        success, message = create_single_user(user)
        if success:
            return CreateUserResponse(
                success=True,
                message=message,
                user=user
            )
        else:
            return CreateUserResponse(
                success=False,
                message=message
            )
    except Exception as e:
        return CreateUserResponse(
            success=False,
            message=f"Error creating user: {str(e)}"
        )

@app.get("/admin/settings", response_model=Settings)
async def get_settings_endpoint(request: Request):
    AuthMiddleware.require_admin(request)
    settings_dict = get_settings()
    return Settings(**settings_dict)

@app.put("/admin/settings", response_model=SettingsResponse)
async def update_settings_endpoint(request: Request, settings_update: SettingsUpdate):
    AuthMiddleware.require_admin(request)
    try:
        # Convert to dict, excluding None values
        update_dict = {k: v for k, v in settings_update.dict().items() if v is not None}
        
        if not update_dict:
            return SettingsResponse(
                success=False,
                message="No settings provided to update"
            )
        
        success = update_settings(update_dict)
        
        if success:
            updated_settings = get_settings()
            return SettingsResponse(
                success=True,
                message="Settings updated successfully",
                settings=Settings(**updated_settings)
            )
        else:
            return SettingsResponse(
                success=False,
                message="Failed to update settings"
            )
    except Exception as e:
        return SettingsResponse(
            success=False,
            message=f"Error updating settings: {str(e)}"
        )

@app.post("/admin/upload-background")
async def upload_background(request: Request, file: UploadFile = File(...)):
    AuthMiddleware.require_admin(request)
    try:
        if not file.content_type or not file.content_type.startswith('image/'):
            return {"success": False, "message": "Please upload an image file"}
        
        # Create uploads directory if it doesn't exist
        import os
        uploads_dir = "static/uploads"
        os.makedirs(uploads_dir, exist_ok=True)
        
        # Save file with unique name
        import uuid
        file_extension = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
        filename = f"background_{uuid.uuid4().hex}.{file_extension}"
        file_path = f"{uploads_dir}/{filename}"
        
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Update settings with new background image path
        web_path = f"/static/uploads/{filename}"
        update_settings({"background_image": web_path})
        
        return {"success": True, "message": "Background image uploaded successfully", "path": web_path}
    
    except Exception as e:
        return {"success": False, "message": f"Error uploading image: {str(e)}"}

@app.delete("/admin/remove-background")
async def remove_background(request: Request):
    AuthMiddleware.require_admin(request)
    try:
        # Get current background image path
        settings = get_settings()
        current_bg = settings.get('background_image', '')
        
        # Remove from settings first
        success = update_settings({"background_image": ""})
        
        if not success:
            return {"success": False, "message": "Failed to update settings"}
        
        # Delete the physical file if it exists and is in uploads folder
        if current_bg and current_bg.startswith('/static/uploads/'):
            import os
            # Convert web path to file path
            file_path = current_bg.replace('/static/', 'static/')
            
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except OSError as e:
                    # File deletion failed but settings were updated
                    return {"success": True, "message": "Background removed but file deletion failed", "warning": str(e)}
        
        return {"success": True, "message": "Background image removed successfully"}
    
    except Exception as e:
        return {"success": False, "message": f"Error removing background: {str(e)}"}

@app.post("/admin/upload-sound")
async def upload_sound(request: Request, sound_type: str = Form(...), file: UploadFile = File(...)):
    AuthMiddleware.require_admin(request)
    try:
        if sound_type not in ['success', 'error']:
            return {"success": False, "message": "Invalid sound type. Must be 'success' or 'error'"}
        
        if not file.content_type or not file.content_type.startswith('audio/'):
            return {"success": False, "message": "Please upload an audio file"}
        
        # Create uploads directory if it doesn't exist
        import os
        uploads_dir = "static/uploads"
        os.makedirs(uploads_dir, exist_ok=True)
        
        # Save file with unique name
        import uuid
        file_extension = file.filename.split('.')[-1] if '.' in file.filename else 'mp3'
        filename = f"{sound_type}_sound_{uuid.uuid4().hex}.{file_extension}"
        file_path = f"{uploads_dir}/{filename}"
        
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Update settings with new sound path
        web_path = f"/static/uploads/{filename}"
        setting_key = f"{sound_type}_sound"
        update_settings({setting_key: web_path})
        
        return {"success": True, "message": f"{sound_type.title()} sound uploaded successfully", "path": web_path}
    
    except Exception as e:
        return {"success": False, "message": f"Error uploading sound: {str(e)}"}

@app.delete("/admin/remove-sound")
async def remove_sound(request: Request, sound_type: str = Form(...)):
    AuthMiddleware.require_admin(request)
    try:
        if sound_type not in ['success', 'error']:
            return {"success": False, "message": "Invalid sound type. Must be 'success' or 'error'"}
        
        # Get current sound path
        settings = get_settings()
        setting_key = f"{sound_type}_sound"
        current_sound = settings.get(setting_key, '')
        
        # Remove from settings first
        success = update_settings({setting_key: ""})
        
        if not success:
            return {"success": False, "message": "Failed to update settings"}
        
        # Delete the physical file if it exists and is in uploads folder
        if current_sound and current_sound.startswith('/static/uploads/'):
            import os
            # Convert web path to file path
            file_path = current_sound.replace('/static/', 'static/')
            
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except OSError as e:
                    # File deletion failed but settings were updated
                    return {"success": True, "message": f"{sound_type.title()} sound removed but file deletion failed", "warning": str(e)}
        
        return {"success": True, "message": f"{sound_type.title()} sound removed successfully"}
    
    except Exception as e:
        return {"success": False, "message": f"Error removing {sound_type} sound: {str(e)}"}

@app.delete("/admin/clear-history")
async def clear_checkin_history_endpoint(request: Request):
    AuthMiddleware.require_admin(request)
    try:
        deleted_count = clear_checkin_history()
        return {
            "success": True,
            "deleted": deleted_count,
            "message": f"Successfully cleared {deleted_count} checkin records"
        }
    except Exception as e:
        return {
            "success": False,
            "message": f"Error clearing checkin history: {str(e)}"
        }

@app.post("/admin/checkin/{employee_id}")
async def manual_checkin(request: Request, employee_id: str):
    AuthMiddleware.require_admin(request)
    try:
        user = get_user_by_employee_id(employee_id)
        
        if not user:
            return {
                "success": False,
                "message": "User not found"
            }
        
        success = create_checkin(employee_id)
        
        if success:
            return {
                "success": True,
                "message": f"Successfully checked in {user.first_name} {user.last_name}",
                "user": {
                    "name": f"{user.first_name} {user.last_name}",
                    "employee_id": employee_id,
                    "table_number": user.table_number
                }
            }
        else:
            return {
                "success": False,
                "message": "Checkin failed"
            }
    except Exception as e:
        return {
            "success": False,
            "message": f"Error during manual checkin: {str(e)}"
        }

@app.delete("/admin/checkout/{employee_id}")
async def manual_checkout(request: Request, employee_id: str):
    AuthMiddleware.require_admin(request)
    try:
        user = get_user_by_employee_id(employee_id)
        
        if not user:
            return {
                "success": False,
                "message": "User not found"
            }
        
        success = checkout_user(employee_id)
        
        if success:
            return {
                "success": True,
                "message": f"Successfully checked out {user.first_name} {user.last_name}",
                "user": {
                    "name": f"{user.first_name} {user.last_name}",
                    "employee_id": employee_id,
                    "table_number": user.table_number
                }
            }
        else:
            return {
                "success": False,
                "message": "No checkin record found to remove"
            }
    except Exception as e:
        return {
            "success": False,
            "message": f"Error during checkout: {str(e)}"
        }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=True)